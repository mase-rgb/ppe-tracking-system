# =============================================================
# Script 3: report.py
# Pulls data from MySQL and generates a formatted Excel report
# with multiple sheets — one per report type.
#
# HOW TO RUN:
#   cd C:\ppe_project
#   python report.py
#
# OUTPUT:
#   C:\ppe_project\PPE_Report_YYYY-MM-DD.xlsx
# =============================================================

import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import config

def get_connection():
    return mysql.connector.connect(**config.DB_CONFIG)

# ── Pull data from MySQL ──────────────────────────────────────
def fetch_all_data(conn):
    # Full status view
    full = pd.read_sql("""
        SELECT
            employee, designation, ppe_type, brand_model,
            issuance_date, expiry_date, source,
            condition_status, current_status
        FROM current_ppe_status
        ORDER BY employee, ppe_type
    """, conn)

    # Needs action
    action = pd.read_sql("""
        SELECT employee, designation, ppe_type, condition_status, issuance_date
        FROM needs_action
    """, conn)

    # Expiring soon
    expiring = pd.read_sql("""
        SELECT employee, designation, ppe_type, issuance_date,
               expiry_date, days_remaining
        FROM expiring_soon
    """, conn)

    # Status summary by employee
    summary = pd.read_sql("""
        SELECT
            employee,
            designation,
            SUM(current_status = 'Active')                 AS active,
            SUM(current_status = 'Expiring Soon')          AS expiring_soon,
            SUM(current_status IN (
                'For Replacement','Damaged',
                'Lost - For Replacement',
                'Misplaced - For Replacement'))            AS needs_action,
            SUM(current_status = 'Expired')                AS expired,
            SUM(current_status IN ('Returned','Replaced')) AS returned_replaced,
            COUNT(*)                                       AS total
        FROM current_ppe_status
        GROUP BY employee, designation
        ORDER BY employee
    """, conn)

    # KPI totals
    kpi = pd.read_sql("""
        SELECT
            SUM(current_status = 'Active')                 AS active,
            SUM(current_status = 'Expiring Soon')          AS expiring_soon,
            SUM(current_status IN (
                'For Replacement','Damaged',
                'Lost - For Replacement',
                'Misplaced - For Replacement'))            AS needs_action,
            SUM(current_status = 'Expired')                AS expired,
            SUM(current_status IN ('Returned','Replaced')) AS returned_replaced,
            COUNT(*)                                       AS total_records
        FROM current_ppe_status
    """, conn)

    # Status log history
    history = pd.read_sql("""
        SELECT
            e.full_name  AS employee,
            c.ppe_type,
            l.old_condition,
            l.new_condition,
            l.old_status,
            l.new_status,
            l.changed_at,
            l.notes
        FROM ppe_status_log l
        JOIN ppe_issuances  i ON i.issuance_id = l.issuance_id
        JOIN employees      e ON e.employee_id = i.employee_id
        JOIN ppe_catalog    c ON c.catalog_id  = i.catalog_id
        ORDER BY l.changed_at DESC
    """, conn)

    return full, action, expiring, summary, kpi, history

# ── Style helpers ─────────────────────────────────────────────
NAVY     = "1E293B"
AMBER    = "F59E0B"
WHITE    = "FFFFFF"
LIGHT    = "F8FAFC"
ALT      = "F1F5F9"
BORDER   = "E2E8F0"

C_GREEN_F = "D1FAE5"; C_GREEN_T = "065F46"
C_RED_F   = "FEE2E2"; C_RED_T   = "991B1B"
C_AMB_F   = "FEF3C7"; C_AMB_T   = "92400E"
C_GRAY_F  = "E2E8F0"; C_GRAY_T  = "334155"
C_BLUE_F  = "DBEAFE"; C_BLUE_T  = "1E40AF"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, color="334155", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def bot(): return Border(bottom=Side(style="thin", color=BORDER))
def thin():
    s = Side(style="thin", color=BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def write_sheet_header(ws, title, col_count):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
    c = ws["A1"]
    c.value = title
    c.fill  = fill(NAVY)
    c.font  = fnt(bold=True, color=WHITE, size=13)
    c.alignment = ctr()

def write_headers(ws, headers, col_widths, row=2):
    ws.row_dimensions[row].height = 20
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill(AMBER); c.font = fnt(bold=True, color=NAVY, size=10)
        c.alignment = ctr(); c.border = thin()

def write_df(ws, df, start_row=3):
    for r_idx, row in enumerate(df.itertuples(index=False), start_row):
        ws.row_dimensions[r_idx].height = 16
        bg = ALT if r_idx % 2 == 0 else LIGHT
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx,
                        value=str(val) if pd.notna(val) else "")
            c.fill = fill(bg); c.font = fnt(size=10)
            c.alignment = lft(); c.border = bot()

def color_status_col(ws, col_idx, start_row, end_row):
    status_colors = {
        "Active":                    (C_GREEN_F, C_GREEN_T),
        "Expiring Soon":             (C_AMB_F,   C_AMB_T),
        "For Replacement":           (C_RED_F,   C_RED_T),
        "Damaged":                   ("FFEDD5",  "7C2D12"),
        "Lost - For Replacement":    ("EDE9FE",  "4C1D95"),
        "Misplaced - For Replacement":("EDE9FE", "4C1D95"),
        "Expired":                   (C_GRAY_F,  C_GRAY_T),
        "Returned":                  (C_BLUE_F,  C_BLUE_T),
        "Replaced":                  (C_BLUE_F,  C_BLUE_T),
    }
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        val  = cell.value or ""
        if val in status_colors:
            fg, ft = status_colors[val]
            cell.fill = fill(fg)
            cell.font = fnt(bold=True, color=ft, size=10)

# ── Build the Excel report ────────────────────────────────────
def build_report(full, action, expiring, summary, kpi, history):
    date_str  = datetime.now().strftime("%Y-%m-%d")
    filename  = f"C:\\ppe_project\\PPE_Report_{date_str}.xlsx"

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    # ── Sheet 1: KPI Summary ──────────────────────────────────
    kpi_ws = writer.book.create_sheet("KPI Summary")
    write_sheet_header(kpi_ws, f"PPE KPI SUMMARY — {date_str}", 2)
    kpi_ws.column_dimensions["A"].width = 28
    kpi_ws.column_dimensions["B"].width = 14

    kpi_headers = ["Status", "Count"]
    kpi_data = [
        ("Active",             int(kpi["active"].iloc[0])),
        ("Expiring Soon",      int(kpi["expiring_soon"].iloc[0])),
        ("Needs Action",       int(kpi["needs_action"].iloc[0])),
        ("Expired",            int(kpi["expired"].iloc[0])),
        ("Returned/Replaced",  int(kpi["returned_replaced"].iloc[0])),
        ("Total Records",      int(kpi["total_records"].iloc[0])),
    ]
    kpi_colors = [C_GREEN_F, C_AMB_F, C_RED_F, C_GRAY_F, C_BLUE_F, "F1F5F9"]
    kpi_txt    = [C_GREEN_T, C_AMB_T, C_RED_T, C_GRAY_T, C_BLUE_T, "334155"]

    for i, (h, w) in enumerate(zip(kpi_headers, [28, 14]), 1):
        kpi_ws.column_dimensions[get_column_letter(i)].width = w
        c = kpi_ws.cell(row=2, column=i, value=h)
        c.fill = fill(AMBER); c.font = fnt(bold=True, color=NAVY, size=10)
        c.alignment = ctr(); c.border = thin()

    for r, ((label, count), fg, ft) in enumerate(zip(kpi_data, kpi_colors, kpi_txt), 3):
        kpi_ws.row_dimensions[r].height = 22
        c1 = kpi_ws.cell(row=r, column=1, value=label)
        c1.fill = fill(fg); c1.font = fnt(bold=True, color=ft, size=11)
        c1.alignment = lft(); c1.border = bot()
        c2 = kpi_ws.cell(row=r, column=2, value=count)
        c2.fill = fill(fg); c2.font = fnt(bold=True, color=ft, size=14)
        c2.alignment = ctr(); c2.border = bot()

    # ── Sheet 2: All PPE Records ──────────────────────────────
    all_ws = writer.book.create_sheet("All PPE Records")
    headers   = ["Employee","Designation","PPE Type","Brand/Model",
                 "Issuance Date","Expiry Date","Source","Condition","Status"]
    col_widths = [18,12,20,16,14,14,9,18,26]
    write_sheet_header(all_ws, f"ALL PPE RECORDS — {date_str}", len(headers))
    write_headers(all_ws, headers, col_widths)
    write_df(all_ws, full)
    color_status_col(all_ws, 9, 3, len(full) + 2)

    # ── Sheet 3: Needs Action ─────────────────────────────────
    act_ws = writer.book.create_sheet("Needs Action")
    act_headers   = ["Employee","Designation","PPE Type","Condition","Issuance Date"]
    act_col_widths = [18,12,20,20,14]
    write_sheet_header(act_ws, f"NEEDS ACTION — {date_str}", len(act_headers))
    write_headers(act_ws, act_headers, act_col_widths)
    write_df(act_ws, action)

    # ── Sheet 4: Expiring Soon ────────────────────────────────
    exp_ws = writer.book.create_sheet("Expiring Soon")
    exp_headers   = ["Employee","Designation","PPE Type","Issuance Date","Expiry Date","Days Remaining"]
    exp_col_widths = [18,12,20,14,14,14]
    write_sheet_header(exp_ws, f"EXPIRING SOON (≤30 days) — {date_str}", len(exp_headers))
    write_headers(exp_ws, exp_headers, exp_col_widths)
    write_df(exp_ws, expiring)

    # ── Sheet 5: Summary by Employee ─────────────────────────
    sum_ws = writer.book.create_sheet("By Employee")
    sum_headers   = ["Employee","Designation","Active","Expiring Soon",
                     "Needs Action","Expired","Returned/Replaced","Total"]
    sum_col_widths = [18,12,10,14,14,10,18,10]
    write_sheet_header(sum_ws, f"PPE SUMMARY BY EMPLOYEE — {date_str}", len(sum_headers))
    write_headers(sum_ws, sum_headers, sum_col_widths)
    write_df(sum_ws, summary)

    # ── Sheet 6: Status Change History ───────────────────────
    hist_ws = writer.book.create_sheet("Status History")
    hist_headers   = ["Employee","PPE Type","Old Condition","New Condition",
                      "Old Status","New Status","Changed At","Notes"]
    hist_col_widths = [18,20,18,18,24,24,20,24]
    write_sheet_header(hist_ws, f"STATUS CHANGE HISTORY — {date_str}", len(hist_headers))
    write_headers(hist_ws, hist_headers, hist_col_widths)
    write_df(hist_ws, history)

    # Remove default sheet
    if "Sheet" in writer.book.sheetnames:
        del writer.book["Sheet"]

    # Set tab colors
    tab_colors = {
        "KPI Summary":      NAVY,
        "All PPE Records":  "059669",
        "Needs Action":     "DC2626",
        "Expiring Soon":    "D97706",
        "By Employee":      "2563EB",
        "Status History":   "4C1D95",
    }
    for name, color in tab_colors.items():
        if name in writer.book.sheetnames:
            writer.book[name].sheet_properties.tabColor = color

    writer.book.save(filename)
    print(f"✓ Report saved: {filename}")
    return filename

# ── Main ──────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 45)
    print("PPE REPORT GENERATOR")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 45)

    try:
        conn = get_connection()
        print("✓ Connected to MySQL")
        full, action, expiring, summary, kpi, history = fetch_all_data(conn)
        conn.close()
        print(f"✓ Data fetched — {len(full)} total records")
        build_report(full, action, expiring, summary, kpi, history)
        print("\n✓ Done. Open the report file to view your PPE data.")
    except mysql.connector.Error as e:
        print(f"✗ Database error: {e}")
    except Exception as e:
        print(f"✗ Error: {e}")
